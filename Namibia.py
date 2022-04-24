import csv
import openpyxl
from openpyxl import Workbook
from openpyxl.styles.colors import Color
from openpyxl.styles import PatternFill
from datetime import date
from datetime import *
import dateutil.relativedelta as REL
import re

def ExecuteNamibia(fileName):
  wb = Workbook()
  ws = wb.active
  ws.title = "Namibia Commission Payout"
  sumCom = 0

  #open csv file and move data over to workbook (xlsx file)
  with open(fileName, encoding="utf8", mode='r') as csv_file:
      csv_reader = csv.DictReader(csv_file)
      #get each row in the csv file
      countRow=2
      for row in csv_reader:
        count=1
        #get each key (column header) from the row dictionary
        #row 1 = column headers (key)
        #all rows after row 1 = values of key (column headers)
        countCol=1
        for key in row:
          if countRow==2:
            ws.cell(row=1, column=countCol, value=key)
            cellValue = row.get(key)
            if countCol==1:
              cellValue = int(cellValue)
            if len(row) > 13:
              if count > 13:
                for i in cellValue:
                  ws.cell(row=countRow, column=countCol, value=i)
              else:
                ws.cell(row=2, column=countCol, value=cellValue)
            else:
              ws.cell(row=2, column=countCol, value=cellValue)
            countCol = countCol + 1
            count += 1
          elif len(row) > 13:
            cellValue = row.get(key)
            if count > 13:
              for i in cellValue:
                ws.cell(row=countRow, column=countCol, value=i)
                countCol = countCol + 1
            else:
              if countCol==1:
                cellValue = int(cellValue)
              ws.cell(row=countRow, column=countCol, value=cellValue)
              countCol = countCol + 1
            count += 1
          else:
            cellValue = row.get(key)
            if countCol==1:
              cellValue = int(cellValue)
            ws.cell(row=countRow, column=countCol, value=cellValue)
            countCol = countCol + 1
        #next row
        countRow = countRow + 1


  
  #iterate through each row of the sheet, skips first row (header)
  # there are 13 columns in the sheet
  rowIndex=2
  bankNamesandSwift = {"First National Bank":["FIRNNANX"], "First National Bank of Namibia Limited":["FIRNNANX"],
                       "Standard Bank Of Namibia Limited":["SBNMNANX"]}
  today = date.today()
  rd = REL.relativedelta(days=1, weekday=REL.FR)
  next_friday = today + rd
  today = next_friday.strftime("%#m-%#d-%Y")
  file = open("Namibia_change_log_" + today + ".txt", "a")
  
  for row in ws.iter_rows(min_row=2):
    username = row[0]
    firstName = row[1]
    lastName = row[2]
    business = row[3]
    taxNum = row[4]
    bankName = row[5]
    routing = row[6]
    upperBool = routing.value.isupper()
    if not upperBool:
      upperVal = routing.value.upper()
      routing.value = upperVal
    account = row[7]
    bankCode = row[8]
    bankCity = row[9]
    bankCountry = row[10]
    amount = row[11]
    currency = row[12]

    if len(row) > 13:
      nextValue = row[13]
      if nextValue.value != None:
        username.fill = PatternFill(fill_type='solid', start_color='f8ff94', end_color='f8ff94')
        firstName.fill = PatternFill(fill_type='solid', start_color='f8ff94', end_color='f8ff94')
        lastName.fill = PatternFill(fill_type='solid', start_color='f8ff94', end_color='f8ff94')
        business.fill = PatternFill(fill_type='solid', start_color='f8ff94', end_color='f8ff94')
        taxNum.fill = PatternFill(fill_type='solid', start_color='f8ff94', end_color='f8ff94')
        bankName.fill = PatternFill(fill_type='solid', start_color='f8ff94', end_color='f8ff94')
        routing.fill = PatternFill(fill_type='solid', start_color='f8ff94', end_color='f8ff94')
        account.fill = PatternFill(fill_type='solid', start_color='f8ff94', end_color='f8ff94')
        bankCode.fill = PatternFill(fill_type='solid', start_color='f8ff94', end_color='f8ff94')
        bankCity.fill = PatternFill(fill_type='solid', start_color='f8ff94', end_color='f8ff94')
        bankCountry.fill = PatternFill(fill_type='solid', start_color='f8ff94', end_color='f8ff94')
        amount.fill = PatternFill(fill_type='solid', start_color='f8ff94', end_color='f8ff94')
        currency.fill = PatternFill(fill_type='solid', start_color='f8ff94', end_color='f8ff94')
        nextValue.fill = PatternFill(fill_type='solid', start_color='f8ff94', end_color='f8ff94')
        rowIndex = rowIndex+1
        continue

    amount.value = float(amount.value)
    sumCom = sumCom + amount.value


    colIndex=1
    #iterate through each cell of the row
    for cell in row:
      if rowIndex == 1:
        continue

      else:
      
        if colIndex==1:
          pass

        #cleans up first name
        if colIndex==2:
          value = cell.value
          list1 = value.split()
          name = ""
          count = 1
          listLength = len(list1)
          for item in list1:
            if count==listLength:
              name = name + item.lower().capitalize()
            else:
              name = name + item.lower().capitalize() + " "
            count += 1
          cell.value = name
          if value != name:
            file.write("ID: " + str(username.value) + ", first name was changed to " + name + " from ''" + value + "''"  +"\n")
          
        #cleans up last name  
        if colIndex==3:
          value = cell.value
          list1 = value.split()
          name = ""
          count = 1
          listLength = len(list1)
          for item in list1:
            if count==listLength:
              name = name + item.lower().capitalize()
            else:
              name = name + item.lower().capitalize() + " "
            count += 1
          cell.value = name
          if value != name:
            file.write("ID: " + str(username.value) + ", last name was changed to " + name + " from ''" + value + "''"  +"\n")

        #set all business names to yellow
        if colIndex==4:
          if cell.value != "":
            cell.fill = PatternFill(fill_type='solid', start_color='f8ff94', end_color='f8ff94')


        #tax number
        if colIndex==5:
          pass

        #bank name
        if colIndex==6:
          #remove extra spaces from bank name and swift
          bName = bankName.value
          bSwift = routing.value
          list1 = bName.split()
          list2 = bSwift.split()
          name = ""
          swift = ""
          count1 = 1
          count2 = 1
          listLength1 = len(list1)
          listLength2 = len(list2)
          #remove spaces from bank name
          for item in list1:
            if count1==listLength1:
              name = name + item
            else:
              name = name + item + " "
            count1 += 1
          bankName.value = name
          if bName != name:
            file.write("ID: " + str(username.value) + ", bank name was changed to " + name + " from ''" + bName + "''"  +"\n")
          #remove spaces from swift
          for item in list2:
            swift = swift + item
            count2 += 1
          routing.value = swift
          if bSwift != swift:
            file.write("ID: " + str(username.value) + ", Swift was changed to " + swift + " from ''" + bSwift + "''"  +"\n")
          #PATH 1
          #if cell value bank name is NOT in the dictionary, change bank name to match swift
          # code if swift code is in dictionary, otherwise mark yellow
          if str(bankName.value) not in bankNamesandSwift:
            count = 0
            for key in bankNamesandSwift:
              if count == 0:
                codes = bankNamesandSwift.get(key)
                for item in codes:
                  #if swift code exists in dictionary
                  if routing.value == item:
                    oldName = bankName.value
                    bankName.value = key
                    file.write("ID: " + str(username.value) + ", bank name changed to " + bankName.value + " from ''" + oldName + "''"  +"\n")
                    count += 1
                    break
              else:
                break
            #if swift code does not exist in dictionary
            if count == 0:
              cell.fill = PatternFill(fill_type='solid', start_color='f8ff94', end_color='f8ff94')
              bankCode.fill = PatternFill(fill_type='solid', start_color='f8ff94', end_color='f8ff94')
              routing.fill = PatternFill(fill_type='solid', start_color='f8ff94', end_color='f8ff94')
              file.write("** ID: " + str(username.value) + ", bank name and routing number highlighted yellow (neither in dictionary)"+"\n")
          #PATH 2
          #if cell value bank name IS in dictionary. If swift code is not in
          # dictionary, change swift to match name. If swift code is in dictionary
          # but they do not match, mark yellow (since we don't know which one is correct)
          elif bankName.value in bankNamesandSwift:
            lyst = bankNamesandSwift.get(bankName.value)
            exists = 0
            exists2 = 0
            for i in lyst:
              if i == routing.value:
                exists += 1
            #if bank name and swift don't match
            if exists == 0:
              for key in bankNamesandSwift:
                if exists2 == 0:
                  codes = bankNamesandSwift.get(key)
                  for item in codes:
                    #if swift code in dictionary
                    if routing.value == item:
                      exists2 += 1
                      cell.fill = PatternFill(fill_type='solid', start_color='f8ff94', end_color='f8ff94')
                      bankCode.fill = PatternFill(fill_type='solid', start_color='f8ff94', end_color='f8ff94')
                      routing.fill = PatternFill(fill_type='solid', start_color='f8ff94', end_color='f8ff94')
                      file.write("** ID: " + str(username.value) + ", bank name and routing number highlighted yellow (both in dictionary but don't match)"+"\n")
                      break
                else:
                  break
              #if swift code not in dicationary
              if exists2 == 0:
                oldR = routing.value
                lyst = bankNamesandSwift.get(bankName.value)
                routing.value = lyst[0]
                file.write("ID: " + str(username.value) + ", routing # changed to " + routing.value + " from ''" + oldR + "''"  +"\n")
                


        #routing number (swift code)
        #this is taken care of in the bank name section (just above)
        if colIndex==7:
          pass

        #account number
        if colIndex==8:
          if cell.value == "":
            cell.fill = PatternFill(fill_type='solid', start_color='f8ff94', end_color='f8ff94')
          isNum = True
          try:
            int(cell.value)
          except ValueError:
            isNum = False
          if not isNum:
            cell.fill = PatternFill(fill_type='solid', start_color='f8ff94', end_color='f8ff94')


        #bank code (COME BACK TO - same as bank name)
        if colIndex==9:
          if bankName.value != bankCode.value:
            bankCode.value = bankName.value
            file.write("ID: " + str(username.value) + ", bank code changed to match bank name: " + bankCode.value + "\n")


        #bank city
        if colIndex==10:
          pass

        #bank country
        if colIndex==11:
          if cell.value != "NA":
            oldBC = cell.value
            cell.value = "NA"
            file.write("ID: " + str(username.value) + ", bank Country changed to NA" + " from ''" + oldBC + "''"  +"\n")

        #amount
        if colIndex==12:
          pass

        #currency
        if colIndex==13:
          if cell.value != "usd":
            oldC = cell.value
            cell.value = "usd"
            file.write("ID: " + str(username.value) + ", currency changed to usd" + " from ''" + oldC + "''"  +"\n")


      colIndex = colIndex+1
    rowIndex = rowIndex+1

  file.write("\n")
  file.write("SUM: " + str(sumCom))
  wb.save("Namibia Commission Payout File " + today + ".xlsx")
  file.close
