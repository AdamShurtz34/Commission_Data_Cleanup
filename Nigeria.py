import csv
import openpyxl
from openpyxl import Workbook
from openpyxl.styles.colors import Color
from openpyxl.styles import PatternFill
from datetime import date
from datetime import *
import dateutil.relativedelta as REL
import re

def ExecuteNigeria(fileName):
  wb = Workbook()
  ws = wb.active
  ws.title = "Nigeria Commission Payout"
  sumCom = 0

  #open csv file and move data over to workbook (xlsx file)
  with open(fileName, encoding="utf8", mode='r') as csv_file:
      csv_reader = csv.DictReader(csv_file)
      #get each row in the csv file
      countRow=2
      for row in csv_reader:
        #get each key (column header) from the row dictionary
        #row 1 = column headers (key)
        #all rows after row 1 = values of key (column headers)
        countCol=1
        for key in row:
          count=1
          cellValue = row.get(key)
          if countRow==2:
            if countCol == 1:
              ws.cell(row=1, column=countCol, value = key)
              ws.cell(row=2, column=countCol, value = cellValue)

            if countCol == 2:
              ws.cell(row=1, column=countCol, value = key)
              ws.cell(row=2, column=countCol, value=cellValue)

            if countCol == 3:
              ws.cell(row=1, column=countCol, value = key)
              ws.cell(row=2, column=countCol, value = cellValue)

            if countCol == 4:
              ws.cell(row=1, column=countCol, value = key)
              day1 = cellValue[3:5]
              month1 = cellValue[:2]
              year1 = cellValue[6:]
              date1 = day1 + "/" + month1 + "/" + year1
              ws.cell(row=2, column=countCol, value=date1)

            if countCol == 5:
              ws.cell(row=1, column=countCol, value = key)
              ws.cell(row=2, column=countCol, value=int(cellValue))

            if countCol == 6:
              ws.cell(row=1, column=countCol, value = key)
              ws.cell(row=2, column=countCol, value = cellValue)

            if countCol == 7:
              ws.cell(row=1, column=countCol, value = key)
              ws.cell(row=2, column=countCol, value = cellValue)

            if countCol == 8:
              ws.cell(row=1, column=countCol, value = key)
              ws.cell(row=2, column=countCol, value = cellValue)

            #next column
            countCol = countCol + 1
          else:
            if len(row) > 8:
              cellValue = row.get(key)
              if count > 8:
                for i in cellValue:
                  ws.cell(row=countRow, column=countCol, value=i)
                  countCol = countCol + 1
              else:
                if countCol==5:
                  cellValue = int(cellValue)
                  ws.cell(row=countRow, column=countCol, value=cellValue)
                elif countCol==4:
                  day3 = cellValue[3:5]
                  month3 = cellValue[:2]
                  year3 = cellValue[6:]
                  date3 = day3 + "/" + month3 + "/" + year3
                  ws.cell(row=countRow, column=countCol, value=date3)
                else: 
                  ws.cell(row=countRow, column=countCol, value=cellValue)
                countCol = countCol + 1
              count += 1
            elif countCol==5:
              cellValue = int(cellValue)
              ws.cell(row=countRow, column=countCol, value=cellValue)
            elif countCol==4:
              day3 = cellValue[3:5]
              month3 = cellValue[:2]
              year3 = cellValue[6:]
              date3 = day3 + "/" + month3 + "/" + year3
              ws.cell(row=countRow, column=countCol, value=date3)
            else: 
              ws.cell(row=countRow, column=countCol, value=cellValue)
            #next column
            countCol = countCol + 1
        #next row
        countRow = countRow + 1


  
  #iterate through each row of the sheet, skips first row (header)
  # there are 13 columns in the sheet
  rowIndex=2
  bankSwiftandSort = {"ABNGNGLA":["044150149"], "DBLNNGLA":["063150269"], "ECOCNGLA":["050150311"], "FCMBNGLA":["214150018"], #Access Bank, Diamond Bank, Ecobank Nigeria, First City Monument Bank
                      "FIDTNGLA":["070150003"], "FBNINGLA":["011151003"], "FSDHNGLA":["501150000"], "GLOUNGLA":["103150001"], #Fidelity Bank, First Bank of Nigeria, FSDH Merchant Bank, Globus Bank Limited
                      "GTBINGLA":["058152052"], "HBCLNGLA":["030150014"], "JAIZNGLA":["301080020"], "PLNINGLA":["082150004"], #Guaranty Trust Bank, Heritage Banking Company, Jaiz Bank, Keystone Bank
                      "CITINGLA":["023150005"], "PRDTNGLA":["076151006"], "UMPLNGLA":["101150001"], "FIRNNGLA":["502150018"], #Citibank Nigeria, Polaris Bank, Providus bank, Rand Merchant Bank Nigeria
                      "SBICNGLX":["221150014"], "SCBLNGLA":["068150015"], "SBICNGLA":["221150014"], "SBICNGLB":["221150645"], #Stanbic Bank, Standard Chartered Bank Nigeria, Stanbic Bank, Stanbic Bank,
                      "NAMENGLA":["232150333"], "SUTGNGLA":["100150001"], "TAJJNGLA":["302080016"], "TTRUNGLA":["102150001"], #Sterling Bank, Suntrust Bank, Taj Bank, Titan Trust Bank
                      "UBNINGLA":["032154568"], "UNAFNGLA":["033152666", "058152052"], "ICITNGLA":["215153593"],              #Union Bank of Nigeria, United Bank for Africa, Unity Bank
                      "WEMANGLA":["035150103"], "ZEIBNGLA":["057150013"], "BKTRUS33":["076151006"]}                           #Wema Bank, Zenith Bank, Deutsche Bank
  today = date.today()
  rd = REL.relativedelta(days=1, weekday=REL.FR)
  next_friday = today + rd
  today = next_friday.strftime("%#m-%#d-%Y")
  file = open("Nigeria_change_log_" + today + ".txt", "a")

  for row in ws.iter_rows(min_row=2):
    reference = row[0]
    fullName = row[1]
    amount = row[2]
    dueDate = row[3]
    username = row[4]
    account = row[5]
    routing = row[6]
    upperBool = routing.value.isupper()
    if not upperBool:
      upperVal = routing.value.upper()
      routing.value = upperVal
    debit = row[7]

    if len(row) > 8 or len(row) < 8:
      nextValue = row[8]
      if nextValue.value != None:
        reference.fill = PatternFill(fill_type='solid', start_color='f8ff94', end_color='f8ff94')
        fullName.fill = PatternFill(fill_type='solid', start_color='f8ff94', end_color='f8ff94')
        amount.fill = PatternFill(fill_type='solid', start_color='f8ff94', end_color='f8ff94')
        dueDate.fill = PatternFill(fill_type='solid', start_color='f8ff94', end_color='f8ff94')
        username.fill = PatternFill(fill_type='solid', start_color='f8ff94', end_color='f8ff94')
        account.fill = PatternFill(fill_type='solid', start_color='f8ff94', end_color='f8ff94')
        routing.fill = PatternFill(fill_type='solid', start_color='f8ff94', end_color='f8ff94')
        debit.fill = PatternFill(fill_type='solid', start_color='f8ff94', end_color='f8ff94')
        nextValue.fill = PatternFill(fill_type='solid', start_color='f8ff94', end_color='f8ff94')
        rowIndex = rowIndex+1
        continue

    amount.value = float(amount.value)
    sumCom = sumCom + amount.value

    colIndex=1
    #iterate through each cell of the row
    for cell in row:
      if rowIndex == 1:
        continue

      else:
        #reference number
        if colIndex==1:
          if cell.value == "":
            cell.fill = PatternFill(fill_type='solid', start_color='f8ff94', end_color='f8ff94')

        #cleans up full name
        if colIndex==2:
          value = fullName.value
          list1 = value.split()
          name = ""
          count = 1
          listLength = len(list1)
          for item in list1:
            if count==listLength:
              name = name + item.lower().capitalize()
            else:
              name = name + item.lower().capitalize() + " "
            count += 1
          fullName.value = name
          if value != name:
            file.write("ID: " + str(username.value) + ", full name was changed to " + name + " from ''" + value + "''"  +"\n")
          
        #payment amount
        if colIndex==3:
          if cell.value == "":
            cell.fill = PatternFill(fill_type='solid', start_color='f8ff94', end_color='f8ff94')

        #due date
        if colIndex==4:
          if cell.value == "":
            cell.fill = PatternFill(fill_type='solid', start_color='f8ff94', end_color='f8ff94')


        #code/ID/username
        if colIndex==5:
          if cell.value == "":
            cell.fill = PatternFill(fill_type='solid', start_color='f8ff94', end_color='f8ff94')

        #Account number
        if colIndex==6:
          stringLyst = account.value.split(" ")
          newACC = ""
          for item in stringLyst:
            oldACC = account.value
            newACC = newACC + item
            account.value = newACC
          if len(account.value) != 10:
            account.fill = PatternFill(fill_type='solid', start_color='f8ff94', end_color='f8ff94')
          else:
            isNum = True
            try:
              int(account.value)
            except ValueError:
              isNum = False
            if not isNum:
              account.fill = PatternFill(fill_type='solid', start_color='f8ff94', end_color='f8ff94')
          if oldACC != newACC:
            file.write("ID: " + str(username.value) + ", account number was changed to" + newACC + " from ''" + oldACC + "''"  +"\n")


        #routing number (swift to sort code)
        if colIndex==7:
          stringLyst = routing.value.split(" ")
          newR = ""
          for item in stringLyst:
            oldR = routing.value
            newR = newR + item
            routing.value = newR
          if oldR != newR:
            file.write("ID: " + str(username.value) + ", swift code was changed to" + newR + " from ''" + oldR + "''"  +"\n")
          if str(routing.value) in bankSwiftandSort:
            codes = bankSwiftandSort.get(routing.value)
            sort = codes[0]
            routing.value = sort
          elif len(routing.value) != 9:
            routing.fill = PatternFill(fill_type='solid', start_color='f8ff94', end_color='f8ff94')
          else:
            isNum = True
            try:
              int(routing.value)
            except ValueError:
              isNum = False
            if not isNum:
              routing.fill = PatternFill(fill_type='solid', start_color='f8ff94',   end_color='f8ff94')

        #debit
        if colIndex==8:
          pass

      colIndex = colIndex+1
    rowIndex = rowIndex+1

  file.write("\n")
  file.write("SUM: " + str(sumCom))
  wb.save("Nigeria Commission Payout File " + today + ".xlsx")
  file.close
