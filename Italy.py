import csv
import openpyxl
from openpyxl import Workbook
from openpyxl.styles.colors import Color
from openpyxl.styles import PatternFill
from datetime import date
from datetime import *
import dateutil.relativedelta as REL
import re

def ExecuteItaly(fileName):
  wb = Workbook()
  ws = wb.active
  ws.title = "Italy Commission Payout"
  sumCom = 0

  #open csv file and move data over to workbook (xlsx file)
  with open(fileName, encoding="utf8", mode='r') as csv_file:
      csv_reader = csv.DictReader(csv_file)
      #get each row in the csv file
      countRow=2
      for row in csv_reader:
        count=1
        #get each key (column header) from the row dictionary
        #row 1 = column headers (key)
        #all rows after row 1 = values of key (column headers)
        countCol=1
        for key in row:
          cellValue = row.get(key)
          if countRow==2:
            if countCol == 1 or countCol == 14:
              ws.cell(row=1, column=countCol, value=key)
              ws.cell(row=2, column=countCol, value=int(cellValue))  
            else:
              ws.cell(row=1, column=countCol, value=key)
              ws.cell(row=2, column=countCol, value=cellValue)

          elif len(row) > 18:
            cellValue = row.get(key)
            if count > 18:
              for i in cellValue:
                ws.cell(row=countRow, column=countCol, value=i)
                countCol = countCol + 1
            else:
              if countCol==1:
                cellValue = int(cellValue)
              ws.cell(row=countRow, column=countCol, value=cellValue)
              countCol = countCol + 1
            count += 1

          else:
            if countCol == 1 or countCol == 14:
              ws.cell(row=countRow, column=countCol, value=int(cellValue))

            else:
              ws.cell(row=countRow, column=countCol, value=cellValue)

          #next column
          countCol = countCol + 1

        #next row
        countRow = countRow + 1


  
  #iterate through each row of the sheet, skips first row (header)
  # there are 13 columns in the sheet
  rowIndex=2
  today = date.today()
  rd = REL.relativedelta(days=1, weekday=REL.FR)
  next_friday = today + rd
  today = next_friday.strftime("%#m-%#d-%Y")
  today2 = date.today()
  file = open("Italy_change_log_" + today + ".txt", "a")

  for row in ws.iter_rows(min_row=2):
    username = row[0]
    lastName = row[1]
    firstName = row[2]
    address = row[3]
    zipC = row[4]
    city = row[5]
    state = row[6]
    VAT = row[7]
    tax = row[8]
    birthD = row[9]
    birthP = row[10]
    birthS = row[11]
    sex = row[12]
    month = row[13]
    amount = row[14]
    INPS = row[15]
    account = row[16]
    routing = row[17]

    if len(row) > 18 or len(row) < 18:
      nextValue = row[18]
      if nextValue.value != None:
        username.fill = PatternFill(fill_type='solid', start_color='f8ff94', end_color='f8ff94')
        firstName.fill = PatternFill(fill_type='solid', start_color='f8ff94', end_color='f8ff94')
        lastName.fill = PatternFill(fill_type='solid', start_color='f8ff94', end_color='f8ff94')
        address.fill = PatternFill(fill_type='solid', start_color='f8ff94', end_color='f8ff94')
        zipC.fill = PatternFill(fill_type='solid', start_color='f8ff94', end_color='f8ff94')
        city.fill = PatternFill(fill_type='solid', start_color='f8ff94', end_color='f8ff94')
        state.fill = PatternFill(fill_type='solid', start_color='f8ff94', end_color='f8ff94')
        VAT.fill = PatternFill(fill_type='solid', start_color='f8ff94', end_color='f8ff94')
        tax.fill = PatternFill(fill_type='solid', start_color='f8ff94', end_color='f8ff94')
        birthD.fill = PatternFill(fill_type='solid', start_color='f8ff94', end_color='f8ff94')
        birthP.fill = PatternFill(fill_type='solid', start_color='f8ff94', end_color='f8ff94')
        birthS.fill = PatternFill(fill_type='solid', start_color='f8ff94', end_color='f8ff94')
        sex.fill = PatternFill(fill_type='solid', start_color='f8ff94', end_color='f8ff94')
        month.fill = PatternFill(fill_type='solid', start_color='f8ff94', end_color='f8ff94')
        amount.fill = PatternFill(fill_type='solid', start_color='f8ff94', end_color='f8ff94')
        INPS.fill = PatternFill(fill_type='solid', start_color='f8ff94', end_color='f8ff94')
        account.fill = PatternFill(fill_type='solid', start_color='f8ff94', end_color='f8ff94')
        routing.fill = PatternFill(fill_type='solid', start_color='f8ff94', end_color='f8ff94')
        nextValue.fill = PatternFill(fill_type='solid', start_color='f8ff94', end_color='f8ff94')
        rowIndex = rowIndex+1
        continue

    amount.value = float(amount.value)
    sumCom = sumCom + amount.value
    
    
    account = row[16]
    upperBool = account.value.isupper()
    if not upperBool:
      upperVal = account.value.upper()
      account.value = upperVal
      
    routing = row[17]
    upperBool = routing.value.isupper()
    if not upperBool:
      upperVal = routing.value.upper()
      routing.value = upperVal
    

    colIndex=1
    #iterate through each cell of the row
    for cell in row:
      if rowIndex == 1:
        continue

      else:
        if colIndex==1:
          pass

        #cleans up last name
        if colIndex==2:
          value = cell.value
          list1 = value.split()
          name = ""
          count = 1
          listLength = len(list1)
          for item in list1:
            if count==listLength:
              name = name + item.lower().capitalize()
            else:
              name = name + item.lower().capitalize() + " "
            count += 1
          cell.value = name
          if value != name:
            file.write("ID: " + str(username.value) + ", last name was changed to " + name + " from ''" + value + "''"  +"\n")
          
        #cleans up first name  
        if colIndex==3:
          value = cell.value
          list1 = value.split()
          name = ""
          count = 1
          listLength = len(list1)
          for item in list1:
            if count==listLength:
              name = name + item.lower().capitalize()
            else:
              name = name + item.lower().capitalize() + " "
            count += 1
          cell.value = name
          if value != name:
            file.write("ID: " + str(username.value) + ", first name was changed to " + name + " from ''" + value + "''"  +"\n")

        #Address
        if colIndex==4:
          value = cell.value
          list1 = value.split()
          adr = ""
          count = 1
          listLength = len(list1)
          for item in list1:
            if count==listLength:
              adr = adr + item.lower().capitalize()
            else:
              adr = adr + item.lower().capitalize() + " "
            count += 1
          cell.value = adr
          if value != adr:
            file.write("ID: " + str(username.value) + ", address was changed to " + adr + " from ''" + value + "''"  +"\n")

        #zip code
        if colIndex==5:
          pass

        #City
        if colIndex==6:
          value = cell.value
          list1 = value.split()
          city2 = ""
          count = 1
          listLength = len(list1)
          for item in list1:
            if count==listLength:
              city2 = city2 + item.lower().capitalize()
            else:
              city2 = city2 + item.lower().capitalize() + " "
            count += 1
          cell.value = city2
          if value != city2:
            file.write("ID: " + str(username.value) + ", city was changed to " + city2 + " from ''" + value + "''"  +"\n")

        #State
        if colIndex==7:
          pass

        #VAT Number
        if colIndex==8:
          pass


        #Tax ID
        if colIndex==9:
          if cell.value == "":
            cell.fill = PatternFill(fill_type='solid', start_color='f8ff94', end_color='f8ff94')


        #Birth Date
        if colIndex==10:
          pass

        #Birth Place
        if colIndex==11:
          pass

        #Birth State
        if colIndex==12:
          pass

        #Sex
        if colIndex==13:
          pass

        #Commission Month
        if colIndex==14:
          pass

        #Gross Commission
        if colIndex==15:
          pass

        #INPS
        if colIndex==16:
          pass

        #IBAN
        if colIndex==17:
          original = account.value
          lyst = account.value.split(" ")
          stringIBAN = ""
          for i in lyst:
            stringIBAN = stringIBAN + i
          account.value = stringIBAN
          if original != stringIBAN:
            file.write("ID: " + str(username.value) + ", IBAN had a space[s] removed " +  "\n")
          isNum = True
          try:
            int(account.value)
          except ValueError:
            isNum = False
          if isNum == True:
            account.fill = PatternFill(fill_type='solid', start_color='f8ff94', end_color='f8ff94')
          if len(account.value) != 27:
            account.fill = PatternFill(fill_type='solid', start_color='f8ff94', end_color='f8ff94')

        #Swift Code
        if colIndex==18:
          original = routing.value
          lyst = routing.value.split(" ")
          stringSWIFT = ""
          for i in lyst:
            stringSWIFT = stringSWIFT + i
          routing.value = stringSWIFT
          if original != stringSWIFT:
            file.write("ID: " + str(username.value) + ", Swift Code had a space[s] removed " +  "\n")
          isNum = True
          try:
            int(routing.value)
          except ValueError:
            isNum = False
          if isNum == True:
            routing.fill = PatternFill(fill_type='solid', start_color='f8ff94', end_color='f8ff94')
          if len(cell.value)!=8 and len(cell.value)!=11:
            routing.fill = PatternFill(fill_type='solid', start_color='f8ff94', end_color='f8ff94')

      colIndex = colIndex+1
    rowIndex = rowIndex+1

  file.write("\n")
  file.write("SUM: " + str(sumCom))
  wb.save("Italy Commission Payout File " + today + ".xlsx")
  file.close
