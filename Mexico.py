import csv
import openpyxl
from openpyxl import Workbook
from openpyxl.styles.colors import Color
from openpyxl.styles import PatternFill
from datetime import date
from datetime import *
import dateutil.relativedelta as REL
import re

def ExecuteMexico(fileName):
  wb = Workbook()
  ws = wb.active
  ws.title = "Mexico Commission Payout"
  sumCom = 0

  #open csv file and move data over to workbook (xlsx file)
  with open(fileName, encoding="utf8", mode='r') as csv_file:
      csv_reader = csv.DictReader(csv_file)
      #get each row in the csv file
      countRow=2
      for row in csv_reader:
        count=1
        #get each key (column header) from the row dictionary
        #row 1 = column headers (key)
        #all rows after row 1 = values of key (column headers)
        countCol=1
        for key in row:
          if countRow==2:
            ws.cell(row=1, column=countCol, value=key)
            cellValue = row.get(key)
            if countCol==1:
              cellValue = int(cellValue)
            if len(row) > 13:
              if count > 13:
                for i in cellValue:
                  ws.cell(row=countRow, column=countCol, value=i)
              else:
                ws.cell(row=2, column=countCol, value=cellValue)
            else:
              ws.cell(row=2, column=countCol, value=cellValue)
            countCol = countCol + 1
            count += 1
          elif len(row) > 13:
            cellValue = row.get(key)
            if count > 13:
              for i in cellValue:
                ws.cell(row=countRow, column=countCol, value=i)
                countCol = countCol + 1
            else:
              if countCol==1:
                cellValue = int(cellValue)
              ws.cell(row=countRow, column=countCol, value=cellValue)
              countCol = countCol + 1
            count += 1
          else:
            cellValue = row.get(key)
            if countCol==1:
              cellValue = int(cellValue)
            ws.cell(row=countRow, column=countCol, value=cellValue)
            countCol = countCol + 1
        #next row
        countRow = countRow + 1


  
  #iterate through each row of the sheet, skips first row (header)
  # there are 13 columns in the sheet
  rowIndex=2
  bankNamesandSwift = {}
  today = date.today()
  rd = REL.relativedelta(days=1, weekday=REL.FR)
  next_friday = today + rd
  today = next_friday.strftime("%#m-%#d-%Y")
  file = open("Mexico_change_log_" + today + ".txt", "a")
  
  for row in ws.iter_rows(min_row=2):
    username = row[0]
    firstName = row[1]
    lastName = row[2]
    business = row[3]
    taxNum = row[4]
    bankName = row[5]
    routing = row[6]
    upperBool = routing.value.isupper()
    if not upperBool:
      upperVal = routing.value.upper()
      routing.value = upperVal
    account = row[7]
    bankCode = row[8]
    bankCity = row[9]
    bankCountry = row[10]
    amount = row[11]
    currency = row[12]

    if len(row) > 13:
      nextValue = row[13]
      if nextValue.value != None:
        username.fill = PatternFill(fill_type='solid', start_color='f8ff94', end_color='f8ff94')
        firstName.fill = PatternFill(fill_type='solid', start_color='f8ff94', end_color='f8ff94')
        lastName.fill = PatternFill(fill_type='solid', start_color='f8ff94', end_color='f8ff94')
        business.fill = PatternFill(fill_type='solid', start_color='f8ff94', end_color='f8ff94')
        taxNum.fill = PatternFill(fill_type='solid', start_color='f8ff94', end_color='f8ff94')
        bankName.fill = PatternFill(fill_type='solid', start_color='f8ff94', end_color='f8ff94')
        routing.fill = PatternFill(fill_type='solid', start_color='f8ff94', end_color='f8ff94')
        account.fill = PatternFill(fill_type='solid', start_color='f8ff94', end_color='f8ff94')
        bankCode.fill = PatternFill(fill_type='solid', start_color='f8ff94', end_color='f8ff94')
        bankCity.fill = PatternFill(fill_type='solid', start_color='f8ff94', end_color='f8ff94')
        bankCountry.fill = PatternFill(fill_type='solid', start_color='f8ff94', end_color='f8ff94')
        amount.fill = PatternFill(fill_type='solid', start_color='f8ff94', end_color='f8ff94')
        currency.fill = PatternFill(fill_type='solid', start_color='f8ff94', end_color='f8ff94')
        nextValue.fill = PatternFill(fill_type='solid', start_color='f8ff94', end_color='f8ff94')
        rowIndex = rowIndex+1
        continue

    amount.value = float(amount.value)
    sumCom = sumCom + amount.value

    colIndex=1
    #iterate through each cell of the row
    for cell in row:
      if rowIndex == 1:
        continue

      else:
      
        if colIndex==1:
          pass

        #cleans up first name
        if colIndex==2:
          value = cell.value
          list1 = value.split()
          name = ""
          count = 1
          listLength = len(list1)
          for item in list1:
            if count==listLength:
              name = name + item.lower().capitalize()
            else:
              name = name + item.lower().capitalize() + " "
            count += 1
          cell.value = name
          if value != name:
            file.write("ID: " + str(username.value) + ", first name was changed to " + name + " from ''" + value + "''"  +"\n")
          
        #cleans up last name  
        if colIndex==3:
          value = cell.value
          list1 = value.split()
          name = ""
          count = 1
          listLength = len(list1)
          for item in list1:
            if count==listLength:
              name = name + item.lower().capitalize()
            else:
              name = name + item.lower().capitalize() + " "
            count += 1
          cell.value = name
          if value != name:
            file.write("ID: " + str(username.value) + ", last name was changed to " + name + " from ''" + value + "''"  +"\n")

        #set all business names to yellow
        if colIndex==4:
          if cell.value != "":
            cell.fill = PatternFill(fill_type='solid', start_color='f8ff94', end_color='f8ff94')


        #tax number
        if colIndex==5:
          pass

        #bank name
        if colIndex==6:
          if cell.value == "":
            cell.fill = PatternFill(fill_type='solid', start_color='f8ff94', end_color='f8ff94')
            bankCode.fill = PatternFill(fill_type='solid', start_color='f8ff94', end_color='f8ff94')

        #routing number (swift code)
        if colIndex==7:
          if str(cell.value) != "0":
            oldR = routing.value
            routing.value = "0"
            file.write("ID: " + str(username.value) + ", routing # changed to " + routing.value + " from ''" + oldR + "''"  +"\n")
          

        #account number
        if colIndex==8:
          if cell.value == "":
            cell.fill = PatternFill(fill_type='solid', start_color='f8ff94', end_color='f8ff94')
          isNum = True
          try:
            int(cell.value)
          except ValueError:
            isNum = False
          if not isNum:
            cell.fill = PatternFill(fill_type='solid', start_color='f8ff94', end_color='f8ff94')
          if len(cell.value) != 18:
            cell.fill = PatternFill(fill_type='solid', start_color='f8ff94', end_color='f8ff94')


        #bank code (COME BACK TO - same as bank name)
        if colIndex==9:
          if bankName.value != bankCode.value:
            bankCode.value = bankName.value
            file.write("ID: " + str(username.value) + ", bank code changed to match bank name: " + bankCode.value + "\n")


        #bank city
        if colIndex==10:
          pass

        #bank country
        if colIndex==11:
          if cell.value != "MX":
            oldBC = cell.value
            cell.value = "MX"
            file.write("ID: " + str(username.value) + ", bank Country changed to MX" + " from ''" + oldBC + "''"  +"\n")

        #amount
        if colIndex==12:
          pass

        #currency
        if colIndex==13:
          if cell.value != "mxn":
            oldC = cell.value
            cell.value = "mxn"
            file.write("ID: " + str(username.value) + ", currency changed to mxn" + " from ''" + oldC + "''"  +"\n")


      colIndex = colIndex+1
    rowIndex = rowIndex+1


  file.write("\n")
  file.write("SUM: " + str(sumCom))
  wb.save("Mexico Commission Payout File " + today + ".xlsx")
  file.close
