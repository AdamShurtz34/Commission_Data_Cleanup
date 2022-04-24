import csv
import openpyxl
from openpyxl import Workbook
from openpyxl.styles.colors import Color
from openpyxl.styles import PatternFill
from datetime import date
from datetime import *
import dateutil.relativedelta as REL

def ExecuteGhanaOther(fileName):
    wb = Workbook()
    ws = wb.active
    ws.title = "Ghana Other Banks"
    sumCom = 0

    with open(fileName, encoding="utf8", mode='r') as csv_file:
        csv_reader = csv.DictReader(csv_file)
        #get each row in the csv file
        countRow=2
        for row in csv_reader:
            count = 1
        #get each key (column header) from the row dictionary
        #row 1 = column headers (key)
        #all rows after row 1 = values of key (column headers)
            countCol=1
            for key in row:
                if countRow==2:
                    ws.cell(row=1, column=countCol, value=key)
                    cellValue = row.get(key)
                    if countCol==1:
                        cellValue = int(cellValue)
                    if len(row) > 12:
                        if count > 12:
                            for i in cellValue:
                                ws.cell(row=2, column=countCol, value=i)
                        else:
                            ws.cell(row=2, column=countCol, value=cellValue)
                    else:
                        ws.cell(row=2, column=countCol, value=cellValue)
                    countCol = countCol + 1
                    count += 1
                elif len(row) > 12:
                    cellValue = row.get(key)
                    if count > 12:
                        for i in cellValue:
                            ws.cell(row=countRow, column=countCol, value=i)
                            countCol = countCol + 1
                    else:
                        if countCol==1:
                            cellValue = int(cellValue)
                        ws.cell(row=countRow, column=countCol, value=cellValue)
                        countCol = countCol + 1
                    count += 1
                else:
                    cellValue = row.get(key)
                    if countCol==1:
                        cellValue = int(cellValue)
                    ws.cell(row=countRow, column=countCol, value=cellValue)
                    countCol = countCol + 1
            #next row
            countRow = countRow + 1


    today = date.today()
    rd = REL.relativedelta(days=1, weekday=REL.FR)
    next_friday = today + rd
    today = next_friday.strftime("%#m-%#d-%Y")
    #--------------------------------------------------------------
    file = open("Ghana-Other_change_log_" + today + ".txt", "a")
    rowIndex=0
    bankNamesandSwift = {"Ghana Commercial Bank":["GHCBGHAC"], "Ecobank Ghana":["ECOCGHAC"], "Stanbic Bank Ghana Limited":["SBICGHAC"],
                         "First Atlantic Bank":["FAMCGHAC"], "Fidelity Bank":["FBLIGHAC"], "Fidelity Bank Ghana":["FBLIGHAC"],
                         "Guaranty Trust Bank Ghana":["GTBIGHAC"], "Prudential Bank Limited":['PUBKGHAC', 'PUBKGHACXXX'], "ADB":["ADNTGHAC"],
                         "Adb":["ADNTGHAC"], "AGRICULTURAL DEVELOPMENT BANK (ADB)":["ADNTGHAC"], "Agricultural Development Bank (adb)":["ADNTGHAC"],
                         "Absa":["BARCGHAC"], "Ecobank":["ECOCGHAC"], "Cal Bank":['ACCCGHAC'], "Calbank":['ACCCGHAC', '140605'],
                         "Consolidated Bank Ghana Limited":['UBGHGHAC'], "ABSA":["BARCGHAC"], "Universal Merchant Bank":['MBGHGHAC'],
                         "Zenith Bank":['ZEBLGHAC'], "Fidelity":['FBLIGHAC'], "Prudential Bank":['PUBKGHAC', 'PUBKGHACTMA'],
                         "Ecobank Ghana Limited":["ECOCGHAC"], "Stanbic Bank Ghana":["SBICGHAC"], "Standard Chartered Bank Limited":['SCBLGHAC'],
                         "Standard Chartered Bank":['SCBLGHAC'],"United Bank For Africa Ghana Ltd":["STBGGHAC"], "Fbn Bank Ghana":["INCEGHAC"],
                         "Stanbic":["SBICGHAC"]}
    for row in ws.iter_rows(min_row=1):
        rowIndex += 1
        if rowIndex == 1:
            continue
        userID = row[0]
        name = row[1]
        bankName = row[2]
        bankCity = row[3]
        bankCountry = row[4]
        swift = row[5]
        account = row[6]
        gross = row[7]
        tax = row[8]
        payment = row[9]
        narration = row[10]
        email = row[11]

        if len(row) > 12 or len(row) < 12:
            nextValue = row[12]
            if nextValue.value != None:
                userID.fill = PatternFill(fill_type='solid', start_color='f8ff94', end_color='f8ff94')
                name.fill = PatternFill(fill_type='solid', start_color='f8ff94', end_color='f8ff94')
                bankName.fill = PatternFill(fill_type='solid', start_color='f8ff94', end_color='f8ff94')
                bankCity.fill = PatternFill(fill_type='solid', start_color='f8ff94', end_color='f8ff94')
                bankCountry.fill = PatternFill(fill_type='solid', start_color='f8ff94', end_color='f8ff94')
                swift.fill = PatternFill(fill_type='solid', start_color='f8ff94', end_color='f8ff94')
                account.fill = PatternFill(fill_type='solid', start_color='f8ff94', end_color='f8ff94')
                gross.fill = PatternFill(fill_type='solid', start_color='f8ff94', end_color='f8ff94')
                tax.fill = PatternFill(fill_type='solid', start_color='f8ff94', end_color='f8ff94')
                payment.fill = PatternFill(fill_type='solid', start_color='f8ff94', end_color='f8ff94')
                narration.fill = PatternFill(fill_type='solid', start_color='f8ff94', end_color='f8ff94')
                email.fill = PatternFill(fill_type='solid', start_color='f8ff94', end_color='f8ff94')
                nextValue.fill = PatternFill(fill_type='solid', start_color='f8ff94', end_color='f8ff94')
                rowIndex = rowIndex+1
                continue

        gross.value = float(gross.value)
        sumCom = sumCom + gross.value

        #clean name
        value = name.value
        list1 = value.split()
        name2 = ""
        count = 1
        listLength = len(list1)
        for item in list1:
            if count==listLength:
                name2 = name2 + item.lower().capitalize()
            else:
                name2 = name2 + item.lower().capitalize() + " "
            count += 1
        name.value = name2
        if value != name2:
            file.write("ID: " + str(userID.value) + ", name was changed to " + name2 + " from ''" + value + "''"  +"\n")


        #clean bank name
        value = bankName.value
        list1 = value.split()
        name2 = ""
        count = 1
        listLength = len(list1)
        for item in list1:
            if count==listLength:
                name2 = name2 + item.lower().capitalize()
            else:
                name2 = name2 + item.lower().capitalize() + " "
            count += 1
        bankName.value = name2
        if value != name2:
            file.write("ID: " + str(userID.value) + ", bank name was changed to " + name2 + " from ''" + value + "''"  +"\n")


        #clean bank city
        value = bankCity.value
        list1 = value.split()
        name2 = ""
        count = 1
        listLength = len(list1)
        for item in list1:
            if count==listLength:
                name2 = name2 + item.lower().capitalize()
            else:
                name2 = name2 + item.lower().capitalize() + " "
            count += 1
        bankCity.value = name2
        if value != name2:
            file.write("ID: " + str(userID.value) + ", bank city was changed to " + name2 + " from ''" + value + "''"  +"\n")


        #account number
        if len(account.value) != 13:
            account.fill = PatternFill(fill_type='solid', start_color='f8ff94', end_color='f8ff94')
        isNum = True
        try:
            int(account.value)
        except ValueError:
            isNum = False
        if not isNum:
            account.fill = PatternFill(fill_type='solid', start_color='f8ff94', end_color='f8ff94')



        #BANK NAME AND SWIFT CODE
        bName = bankName.value
        bSwift = swift.value
        list2 = bSwift.split()
        swift2 = ""
        count2 = 1
        listLength2 = len(list2)
        #remove spaces from swift
        for item in list2:
            swift2 = swift2 + item
            count2 += 1
        swift.value = swift2
        if bSwift != swift2:
            file.write("ID: " + str(userID.value) + ", Swift was changed to " + swift2 + " from ''" + bSwift + "''"  +"\n")
        #PATH 1
        #if cell value bank name is NOT in the dictionary, change bank name to match swift
        # code if swift code is in dictionary, otherwise mark yellow
        if str(bankName.value) not in bankNamesandSwift:
            count = 0
            for key in bankNamesandSwift:
                if count == 0:
                    codes = bankNamesandSwift.get(key)
                    for item in codes:
                        #if swift code exists in dictionary
                        if swift.value == item:
                            oldName = bankName.value
                            bankName.value = key
                            file.write("ID: " + str(userID.value) + ", bank name changed to " + bankName.value + " from ''" + oldName + "''"  +"\n")
                            count += 1
                            break
                else:
                    break
            #if swift code does not exist in dictionary
            if count == 0:
                bankName.fill = PatternFill(fill_type='solid', start_color='f8ff94', end_color='f8ff94')
                swift.fill = PatternFill(fill_type='solid', start_color='f8ff94', end_color='f8ff94')
                file.write("** ID: " + str(userID.value) + ", bank name and swift code highlighted yellow (neither in dictionary)"+"\n")
        #PATH 2
        #if cell value bank name IS in dictionary. If swift code is not in
        # dictionary, change swift to match name. If swift code is in dictionary
        # but they do not match, mark yellow (since we don't know which one is correct)
        elif bankName.value in bankNamesandSwift:
            lyst = bankNamesandSwift.get(bankName.value)
            exists = 0
            exists2 = 0
            for i in lyst:
                if i == swift.value:
                    exists += 1
            #if bank name and swift don't match
            if exists == 0:
                for key in bankNamesandSwift:
                    if exists2 == 0:
                        codes = bankNamesandSwift.get(key)
                        for item in codes:
                            #if swift code in dictionary
                            if swift.value == item:
                                exists2 += 1
                                bankName.fill = PatternFill(fill_type='solid', start_color='f8ff94', end_color='f8ff94')
                                swift.fill = PatternFill(fill_type='solid', start_color='f8ff94', end_color='f8ff94')
                                file.write("** ID: " + str(username.value) + ", bank name and swift highlighted yellow (both in dictionary but don't match)"+"\n")
                                break
                    else:
                        break
                #if swift code not in dicationary
                if exists2 == 0:
                    oldR = swift.value
                    lyst = bankNamesandSwift.get(bankName.value)
                    swift.value = lyst[0]
                    file.write("ID: " + str(userID.value) + ", swift code changed to " + swift.value + " from ''" + oldR + "''"  +"\n")


        #bank country
        if bankCountry.value != "GH":
            oldBC = bankCountry.value
            bankCountry.value = "GH"
            file.write("ID: " + str(userID.value) + ", bank Country changed to GH" + " from ''" + oldBC + "''"  +"\n")

        #Gross Commissions
        gross.value = float(gross.value)
        if gross.value == "" or gross.value == None:
            gross.fill = PatternFill(fill_type='solid', start_color='f8ff94', end_color='f8ff94')

        #tax withheld
        tax.value = gross.value * .05
        if tax.value == "" or tax.value == None:
            tax.fill = PatternFill(fill_type='solid', start_color='f8ff94', end_color='f8ff94')

        #payment
        payment.value = gross.value - tax.value
        if payment.value == "" or payment.value == None:
            payment.fill = PatternFill(fill_type='solid', start_color='f8ff94', end_color='f8ff94')

        #narration
        if narration.value == "" or narration.value == None:
            narration.fill = PatternFill(fill_type='solid', start_color='f8ff94', end_color='f8ff94')

        #email
        if email.value == "" or email.value == None:
            email.fill = PatternFill(fill_type='solid', start_color='f8ff94', end_color='f8ff94')


    file.write("\n")
    file.write("SUM: " + str(sumCom))
    wb.save("Ghana Other Banks Commissions " + today + ".xlsx")
    file.close
                










            
        

