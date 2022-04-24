import csv
import openpyxl
from openpyxl import Workbook
from openpyxl.styles.colors import Color
from openpyxl.styles import PatternFill
from datetime import date
from datetime import *
import dateutil.relativedelta as REL

def ExecuteGhanaAccess(fileName):
    wb = Workbook()
    ws = wb.active
    ws.title = "Ghana Access Bank"
    sumCom = 0

    with open(fileName, encoding="utf8", mode='r') as csv_file:
        csv_reader = csv.DictReader(csv_file)
        #get each row in the csv file
        countRow=2
        for row in csv_reader:
            count = 1
        #get each key (column header) from the row dictionary
        #row 1 = column headers (key)
        #all rows after row 1 = values of key (column headers)
            countCol=1
            for key in row:
                if countRow==2:
                    ws.cell(row=1, column=countCol, value=key)
                    cellValue = row.get(key)
                    if countCol==1:
                        cellValue = int(cellValue)
                    if len(row) > 11:
                        if count > 11:
                            for i in cellValue:
                                ws.cell(row=2, column=countCol, value=i)
                        else:
                            ws.cell(row=2, column=countCol, value=cellValue)
                    else:
                        ws.cell(row=2, column=countCol, value=cellValue)
                    countCol = countCol + 1
                    count += 1
                elif len(row) > 11:
                    cellValue = row.get(key)
                    if count > 11:
                        for i in cellValue:
                            ws.cell(row=countRow, column=countCol, value=i)
                            countCol = countCol + 1
                    else:
                        if countCol==1:
                            cellValue = int(cellValue)
                        ws.cell(row=countRow, column=countCol, value=cellValue)
                        countCol = countCol + 1
                    count += 1
                else:
                    cellValue = row.get(key)
                    if countCol==1:
                        cellValue = int(cellValue)
                    ws.cell(row=countRow, column=countCol, value=cellValue)
                    countCol = countCol + 1
            #next row
            countRow = countRow + 1


    today = date.today()
    rd = REL.relativedelta(days=1, weekday=REL.FR)
    next_friday = today + rd
    today = next_friday.strftime("%#m-%#d-%Y")
    #--------------------------------------------------------------
    file = open("Ghana-Access_change_log_" + today + ".txt", "a")
    rowIndex=0
    bankNamesandSwift = {}
    for row in ws.iter_rows(min_row=1):
        rowIndex += 1
        if rowIndex == 1:
            continue
        userID = row[0]
        name = row[1]
        bankName = row[2]
        bankCity = row[3]
        bankCountry = row[4]
        account = row[5]
        gross = row[6]
        tax = row[7]
        payment = row[8]
        narration = row[9]
        email = row[10]

        if len(row) > 11 or len(row) < 11:
            nextValue = row[11]
            if nextValue.value != None:
                userID.fill = PatternFill(fill_type='solid', start_color='f8ff94', end_color='f8ff94')
                name.fill = PatternFill(fill_type='solid', start_color='f8ff94', end_color='f8ff94')
                bankName.fill = PatternFill(fill_type='solid', start_color='f8ff94', end_color='f8ff94')
                bankCity.fill = PatternFill(fill_type='solid', start_color='f8ff94', end_color='f8ff94')
                bankCountry.fill = PatternFill(fill_type='solid', start_color='f8ff94', end_color='f8ff94')
                account.fill = PatternFill(fill_type='solid', start_color='f8ff94', end_color='f8ff94')
                gross.fill = PatternFill(fill_type='solid', start_color='f8ff94', end_color='f8ff94')
                tax.fill = PatternFill(fill_type='solid', start_color='f8ff94', end_color='f8ff94')
                payment.fill = PatternFill(fill_type='solid', start_color='f8ff94', end_color='f8ff94')
                narration.fill = PatternFill(fill_type='solid', start_color='f8ff94', end_color='f8ff94')
                email.fill = PatternFill(fill_type='solid', start_color='f8ff94', end_color='f8ff94')
                nextValue.fill = PatternFill(fill_type='solid', start_color='f8ff94', end_color='f8ff94')
                rowIndex = rowIndex+1
                continue
            
        gross.value = float(gross.value)
        sumCom = sumCom + gross.value

        #clean name
        value = name.value
        list1 = value.split()
        name2 = ""
        count = 1
        listLength = len(list1)
        for item in list1:
            if count==listLength:
                name2 = name2 + item.lower().capitalize()
            else:
                name2 = name2 + item.lower().capitalize() + " "
            count += 1
        name.value = name2
        if value != name2:
            file.write("ID: " + str(userID.value) + ", name was changed to " + name2 + " from ''" + value + "''"  +"\n")


        #clean bank name
        value = bankName.value
        list1 = value.split()
        name2 = ""
        count = 1
        listLength = len(list1)
        for item in list1:
            if count==listLength:
                name2 = name2 + item.lower().capitalize()
            else:
                name2 = name2 + item.lower().capitalize() + " "
            count += 1
        bankName.value = name2
        if value != name2:
            file.write("ID: " + str(userID.value) + ", bank name was changed to " + name2 + " from ''" + value + "''"  +"\n")

        if bankName.value != "Access" and bankName.value != "Access Bank" and bankName.value != "Access Bank Ghana":
            bankName.fill = PatternFill(fill_type='solid', start_color='f8ff94', end_color='f8ff94')


        #clean bank city
        value = bankCity.value
        list1 = value.split()
        name2 = ""
        count = 1
        listLength = len(list1)
        for item in list1:
            if count==listLength:
                name2 = name2 + item.lower().capitalize()
            else:
                name2 = name2 + item.lower().capitalize() + " "
            count += 1
        bankCity.value = name2
        if value != name2:
            file.write("ID: " + str(userID.value) + ", bank city was changed to " + name2 + " from ''" + value + "''"  +"\n")


        #account number
        if len(account.value) != 13:
            account.fill = PatternFill(fill_type='solid', start_color='f8ff94', end_color='f8ff94')
        isNum = True
        try:
            int(account.value)
        except ValueError:
            isNum = False
        if not isNum:
            account.fill = PatternFill(fill_type='solid', start_color='f8ff94', end_color='f8ff94')

        #bank country
        if bankCountry.value != "GH":
            oldBC = bankCountry.value
            bankCountry.value = "GH"
            file.write("ID: " + str(userID.value) + ", bank Country changed to GH" + " from ''" + oldBC + "''"  +"\n")

        #Gross Commissions
        gross.value = float(gross.value)
        if gross.value == "" or gross.value == None:
            gross.fill = PatternFill(fill_type='solid', start_color='f8ff94', end_color='f8ff94')

        #tax withheld
        tax.value = gross.value * .05
        if tax.value == "" or tax.value == None:
            tax.fill = PatternFill(fill_type='solid', start_color='f8ff94', end_color='f8ff94')

        #payment
        payment.value = gross.value - tax.value
        if payment.value == "" or payment.value == None:
            payment.fill = PatternFill(fill_type='solid', start_color='f8ff94', end_color='f8ff94')

        #narration
        if narration.value == "" or narration.value == None:
            narration.fill = PatternFill(fill_type='solid', start_color='f8ff94', end_color='f8ff94')

        #email
        if email.value == "" or email.value == None:
            email.fill = PatternFill(fill_type='solid', start_color='f8ff94', end_color='f8ff94')


    file.write("\n")
    file.write("SUM: " + str(sumCom))
    wb.save("Ghana Access Bank Commissions " + today + ".xlsx")
    file.close
        







            
