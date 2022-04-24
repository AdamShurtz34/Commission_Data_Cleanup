import csv
import openpyxl
from openpyxl import Workbook
from openpyxl.styles.colors import Color
from openpyxl.styles import PatternFill
from datetime import date
from datetime import *
import dateutil.relativedelta as REL
import re

def ExecuteUS(fileName):
  wb = Workbook()
  ws = wb.active
  ws.title = "US Commission Payout"
  sumCom = 0

  #open csv file and move data over to workbook (xlsx file)
  with open(fileName, encoding="utf8", mode='r') as csv_file:
      csv_reader = csv.DictReader(csv_file)
      #get each row in the csv file
      countRow=2
      for row in csv_reader:
        count=1
        #get each key (column header) from the row dictionary
        #row 1 = column headers (key)
        #all rows after row 1 = values of key (column headers)
        countCol=1
        for key in row:
          if countRow==2:
            ws.cell(row=1, column=countCol, value=key)
            cellValue = row.get(key)
            if countCol==1:
              cellValue = int(cellValue)
            if len(row) > 5:
              if count > 5:
                for i in cellValue:
                  ws.cell(row=countRow, column=countCol, value=i)
              else:
                ws.cell(row=2, column=countCol, value=cellValue)
            else:
              ws.cell(row=2, column=countCol, value=cellValue)
            countCol = countCol + 1
            count += 1
          elif len(row) > 5:
            cellValue = row.get(key)
            if count > 5:
              for i in cellValue:
                ws.cell(row=countRow, column=countCol, value=i)
                countCol = countCol + 1
            else:
              if countCol==1:
                cellValue = int(cellValue)
              ws.cell(row=countRow, column=countCol, value=cellValue)
              countCol = countCol + 1
            count += 1
          else:
            cellValue = row.get(key)
            if countCol==1:
              cellValue = int(cellValue)
            ws.cell(row=countRow, column=countCol, value=cellValue)
            countCol = countCol + 1
        #next row
        countRow = countRow + 1


  
  #iterate through each row of the sheet, skips first row (header)
  # there are 13 columns in the sheet
  rowIndex=2
  bankNamesandSwift = {}
  today = date.today()
  rd = REL.relativedelta(days=1, weekday=REL.FR)
  next_friday = today + rd
  file = open("US_change_log_" + next_friday.strftime("%#m-%#d-%Y") + ".txt", "a")

  for row in ws.iter_rows(min_row=2):
    username = row[0]
    fullName = row[1]
    routing = row[2]
    account = row[3]
    amount = row[4]

    if len(row) > 5 or len(row) < 5:
      nextValue = row[5]
      if nextValue.value != None:
        username.fill = PatternFill(fill_type='solid', start_color='f8ff94', end_color='f8ff94')
        fullName.fill = PatternFill(fill_type='solid', start_color='f8ff94', end_color='f8ff94')
        routing.fill = PatternFill(fill_type='solid', start_color='f8ff94', end_color='f8ff94')
        account.fill = PatternFill(fill_type='solid', start_color='f8ff94', end_color='f8ff94')
        amount.fill = PatternFill(fill_type='solid', start_color='f8ff94', end_color='f8ff94')
        nextValue.fill = PatternFill(fill_type='solid', start_color='f8ff94', end_color='f8ff94')
        rowIndex = rowIndex+1
        continue

    amount.value = float(amount.value)
    sumCom = sumCom + amount.value


    colIndex=1
    #iterate through each cell of the row
    for cell in row:
      if rowIndex == 1:
        continue

      else:
        if colIndex==1:
          pass

        #cleans up full name
        if colIndex==2:
          value = fullName.value
          list1 = value.split()
          name = ""
          count = 1
          listLength = len(list1)
          for item in list1:
            if count==listLength:
              if item.isupper():
                cell.fill = PatternFill(fill_type='solid', start_color='f8ff94', end_color='f8ff94')
              if item.islower():
                name = name + item.capitalize()
              else:
                name = name + item
            else:
              if item.islower():
                name = name + item.capitalize() + " "
              else:
                name = name + item + " "
            count += 1
          fullName.value = name
          if value != name:
            file.write("ID: " + str(username.value) + ", full name was changed to " + name + " from ''" + value + "''"  +"\n")


          #check fullname for special characters
          value = fullName.value
          for char in value:
            num = ord(char)
            if (num < 65) or (num > 90):
              if (num < 97) or (num > 122):
                if num!=32 and num!=45 and num!=46:
                  fullName.fill = PatternFill(fill_type='solid', start_color='f8ff94', end_color='f8ff94')
          

        #routing number
        if colIndex==3:
          if len(cell.value)!=9:
            cell.fill = PatternFill(fill_type='solid', start_color='f8ff94', end_color='f8ff94')
          isNum = True
          try:
            int(cell.value)
          except ValueError:
            isNum = False
          if not isNum:
            cell.fill = PatternFill(fill_type='solid', start_color='f8ff94', end_color='f8ff94')

        #account number
        if colIndex==4:
          if cell.value == "":
            cell.fill = PatternFill(fill_type='solid', start_color='f8ff94', end_color='f8ff94')
          isNum = True
          try:
            int(cell.value)
          except ValueError:
            isNum = False
          if not isNum:
            cell.fill = PatternFill(fill_type='solid', start_color='f8ff94', end_color='f8ff94')

        #amount
        if colIndex==5:
          pass

      colIndex = colIndex+1
    rowIndex = rowIndex+1

  file.write("\n")
  file.write("SUM: " + str(sumCom))
  wb.save("US Commission Payout File " + next_friday.strftime("%#m-%#d-%Y") + ".xlsx")
  file.close
